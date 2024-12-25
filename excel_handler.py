import os
import openpyxl as xl
import datetime as dt
import shutil
import logging


MASTER_EXCEL = './excel_sheets/master.xlsx'
current_workbook = f'./excel_sheets/CC Attendance {dt.datetime.now().strftime("%d-%m-%Y")}.xlsx'

def generate_new_attendance_workbook():
    if os.path.exists(current_workbook):
        return 0
    curr = xl.Workbook()
    filename = f'./excel_sheets/CC Attendance {dt.datetime.now().strftime("%d-%m-%Y")}.xlsx'
    curr.save(filename)
    curr.close()
    shutil.copy(MASTER_EXCEL, filename)
    
    return filename.lstrip('./excel_sheets/')


def get_session():
    current_hour = dt.datetime.now().hour
    if current_hour < 12:
        return 'M'
    else:
        return 'E'

def add_attendance(roll_number: str, session: int = 3):
    if not os.path.exists(current_workbook):
        raise "Attendance workbook not found. Please generate a new attendance workbook."
        
    wb = xl.load_workbook(current_workbook)
    marked = False

    sheets = wb.sheetnames
    
    for sheet_name in sheets:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2): 
                if row[2].value == roll_number:
                    if session == 1:
                        ws.cell(row=row[0].row, column=4, value="P")    # Morning column
                        ws.cell(row=row[0].row, column=5, value="A")    # Evening column
                        print('\033[92m'+f"Roll number {roll_number} marked present ONLY for MORNING session."+'\033[0m')
                        marked = True
                        break
                    elif session == 2:
                        ws.cell(row=row[0].row, column=4, value="A")    # Morning column
                        ws.cell(row=row[0].row, column=5, value="P")    #evening column
                        print('\033[92m'+f"Roll number {roll_number} marked present ONLY for AFTERNOON session."+ '\033[0m')
                        marked = True
                        break
                    else:
                        ws.cell(row=row[0].row, column=4, value="P")  # Morning column
                        ws.cell(row=row[0].row, column=5, value="P")  # Evening column
                        print('\033[92m'+f"Roll number {roll_number} marked present for BOTH sessions."+'\033[0m')
                        marked = True
                        break  
    
    if not marked:
        print('\033[91m'+"Roll number not found in the attendance sheet."+ '\033[0m')
        return 0
    
    
    wb.save(current_workbook)
    logging.info(f"Changes saved")
    wb.save(current_workbook)
    wb.close()

    return 1


def mark_absent():
    wb = xl.load_workbook(current_workbook)
    sheets = wb.sheetnames

    for sheet_name in sheets:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2):
            if row[3].value != 'P':
                ws.cell(row=row[0].row, column=4, value="A")
            if row[4].value != 'P':
                ws.cell(row=row[0].row, column=5, value="A")

    wb.save(current_workbook)
    wb.close()
    return "Cleanup complete"

def add_conditional_formatting():
    wb = xl.load_workbook(current_workbook)
    sheets = wb.sheetnames

    for sheet_name in sheets:
        ws = wb[sheet_name]
        
        red_fill = xl.styles.PatternFill(start_color='FFFF9999', end_color='FFFF9999', fill_type='solid')
        green_fill = xl.styles.PatternFill(start_color='FF99FF99', end_color='FF99FF99', fill_type='solid')

        for row in ws.iter_rows(min_row=2):
            if row[3].value == 'A':
                ws.cell(row=row[0].row, column=4).fill = red_fill
            if row[4].value == 'A':
                ws.cell(row=row[0].row, column=5).fill = red_fill
            if row[3].value == 'P':
                ws.cell(row=row[0].row, column=4).fill = green_fill
            if row[4].value == 'P':
                ws.cell(row=row[0].row, column=5).fill = green_fill

    wb.save(current_workbook)
    wb.close()
    return "Conditional formatting added"

if __name__ == '__main__':
    #print(generate_new_attendance_workbook())
    #print(add_attendance('22CE1021'))
    #print(add_attendance('22CE1050'))
    #print(add_attendance('22CB1083'))
    #print(add_attendance('22CB1019'))
    #print(add_attendance('23CE1017'))
    #print(mark_absent())
    #print(add_conditional_formatting())
    pass